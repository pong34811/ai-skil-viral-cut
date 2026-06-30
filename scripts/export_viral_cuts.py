import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
import os, sys, re, argparse

REQUIRED_FIELDS = ["category", "title", "start_time", "end_time", "filename", "duration", "score"]
TIME_RE = re.compile(r"^\d{2}:\d{2}:\d{2}$")

def validate_cuts(cuts):
    errors = []
    for i, cut in enumerate(cuts):
        for field in REQUIRED_FIELDS:
            if field not in cut or not cut[field]:
                errors.append(f"Cut #{i+1}: missing required field '{field}'")
        score = cut.get("score")
        if score is not None:
            try:
                s = float(score)
                if s < 0 or s > 10:
                    errors.append(f"Cut #{i+1}: score {s} out of range 0-10")
            except (ValueError, TypeError):
                errors.append(f"Cut #{i+1}: score '{score}' not a number")
        st = cut.get("start_time", "")
        et = cut.get("end_time", "")
        if st and not TIME_RE.match(st):
            errors.append(f"Cut #{i+1}: start_time '{st}' not HH:MM:SS")
        if et and not TIME_RE.match(et):
            errors.append(f"Cut #{i+1}: end_time '{et}' not HH:MM:SS")
    return errors

def export_viral_cuts_to_excel(cuts, output_path=None):
    if output_path is None:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        output_path = f"viral-cut-{today}.xlsx"
    sorted_cuts = sorted(cuts, key=lambda c: float(c.get("score", 0) or 0), reverse=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viral Cuts"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "#", "ประเภท (Category)", "ชื่อคลิป", "เวลาเริ่มต้น", "เวลาสิ้นสุด",
        "ชื่อไฟล์นำเข้า (.mp4)", "ความยาวคลิป (Duration)", "คะแนน Viral"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, cut in enumerate(sorted_cuts, 2):
        values = [
            row_idx - 1,
            cut.get("category", ""),
            cut.get("title", ""),
            cut.get("start_time", ""),
            cut.get("end_time", ""),
            cut.get("filename", ""),
            cut.get("duration", ""),
            cut.get("score", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    col_widths = [5, 22, 40, 14, 14, 30, 22, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30
    for row in range(2, len(cuts) + 2):
        ws.row_dimensions[row].height = 40

    ws.auto_filter.ref = f"A1:H{len(cuts) + 1}"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path

def export_all(cuts, date_str=None):
    errors = validate_cuts(cuts)
    if errors:
        for e in errors:
            print(f"[Validation Error] {e}")
        print("Export cancelled. Fix errors and retry.")
        return

    if date_str is None:
        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    filename = f"viral-cut-{date_str}"

    os.makedirs(f"outputs/{date_str}", exist_ok=True)
    out = export_viral_cuts_to_excel(cuts, output_path=f"outputs/{date_str}/{filename}.xlsx")
    print(f"Output: {out}")

    ref = export_viral_cuts_to_excel(cuts, output_path=f"references/{date_str}/{filename}.xlsx")
    print(f"Reference: {ref}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export viral cuts to Excel")
    parser.add_argument("--cuts", nargs="*", help="cuts in format: category|title|start|end|filename|duration|score")
    parser.add_argument("--date", help="date string YYYY-MM-DD (default: today UTC)")
    args = parser.parse_args()

    if args.cuts:
        parsed = []
        for c in args.cuts:
            parts = c.split("|")
            if len(parts) >= 7:
                parsed.append({
                    "category": parts[0],
                    "title": parts[1],
                    "start_time": parts[2],
                    "end_time": parts[3],
                    "filename": parts[4],
                    "duration": parts[5],
                    "score": float(parts[6]),
                })
        export_all(parsed, date_str=args.date)
    else:
        print("Usage: python scripts/export_viral_cuts.py --cuts \"Cat|Title|00:00:00|00:01:00|file.mp4|1m|7.5\"")